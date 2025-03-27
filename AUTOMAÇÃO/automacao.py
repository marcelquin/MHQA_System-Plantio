import openpyxl
import asyncio
import requests
from dotenv import load_dotenv
import os
from time import sleep

load_dotenv()

#TELEGRAM_TOKEN = os.getenv("TOKEN")
#CHAT_ID = os.getenv("CHAT_ID")
#     #métodos de automação

async def Adicionar_Area_Palntio():
    tabela = openpyxl.load_workbook("AreaPlantio.xlsx")
    tabela_dados = tabela['Plan1']
    for linha in tabela_dados.iter_rows(min_row=2):
        nomeIdentificador = linha[0].value if linha[0].value else ""
        dimencao = linha[1].value if linha[1].value else ""
        gps = linha[2].value if linha[2].value else ""
        url = os.getenv("AREA_POST")
        params = {
                'nomeIdentificador': nomeIdentificador,
                'dimencao': dimencao,
                'gps': gps
            } 
        Enviar_Informacao(params, url)

async def Adicionar_Subarea_Palntio():
    tabela = openpyxl.load_workbook("SubareaPlantio.xlsx")
    tabela_dados = tabela['Plan1']
    for linha in tabela_dados.iter_rows(min_row=2):
        cor = linha[0].value if linha[0].value else ""
        tamanho = linha[1].value if linha[1].value else ""
        nomeAreaPlantio = linha[2].value if linha[2].value else ""
        quantidade = linha[3].value if linha[3].value else ""
        url = os.getenv("AUTOMACAO_SUBAREA_POST")
        params = {
                'cor': cor,
                'tamanho': tamanho,
                'nomeAreaPlantio': nomeAreaPlantio,
                'quantidade': quantidade
            } 
        Enviar_Informacao(params, url)

async def Alterar_Ciclo():
    tabela = openpyxl.load_workbook("AlterarCiclo.xlsx")
    tabela_dados = tabela['Plan1']
    for linha in tabela_dados.iter_rows(min_row=2):
        codigo = linha[0].value if linha[0].value else ""
        numeroSubarea = linha[1].value if linha[1].value else ""
        faseatual = linha[2].value if linha[0].value else ""
        url = os.getenv("PLANTA_ALTERAR_CICLO")
        print(url)
        params = {
                'codigo': codigo,
                'numeroSubarea': numeroSubarea,
                'faseatual': faseatual
            } 
        Editar_Informacao(params, url)

async def Adicionar_Adubacao_Geral():
    tabela = openpyxl.load_workbook("AdubacaoGeral.xlsx")
    tabela_dados = tabela['Plan1']
    for linha in tabela_dados.iter_rows(min_row=2):
        nomeIdentificador = linha[0].value if linha[0].value else ""
        adubacao = linha[1].value if linha[1].value else ""
        params = {
            'nomeIdentificador': nomeIdentificador,
            'adubacao': adubacao
        }
        URL = os.getenv("AREA_ADUBACAO_GERAL")
        Editar_Informacao(params, URL) 

async def Adicionar_Adubacao_Individual():
    tabela = openpyxl.load_workbook("AdubacaoIndividual.xlsx")
    tabela_dados = tabela['Plan1']
    for linha in tabela_dados.iter_rows(min_row=2):
        nomeAreaPlantio = linha[0].value if linha[0].value else ""
        numeroSubarea = linha[1].value if linha[1].value else ""
        resumoAdubacao = linha[2].value if linha[2].value else ""
        params = {
            'nomeAreaPlantio': nomeAreaPlantio,
            'numeroSubarea': numeroSubarea,
            'resumoAdubacao': resumoAdubacao
        }
        URL = os.getenv("SUBAREA_PUT_ADUBACAO_INDIVIDUAL")
        Editar_Informacao(params, URL)            
            
async def RealizarEnxertia():
    tabela = openpyxl.load_workbook("Enxertia.xlsx")
    tabela_dados = tabela['Plan1']
    for linha in tabela_dados.iter_rows(min_row=2):
        codigoPlantaDoadora = linha[0].value if linha[0].value else ""
        codigoPlantaReceptora = linha[1].value if linha[1].value else ""
        params = {
            'codigoPlantaDoadora': codigoPlantaDoadora,
            'codigoPlantaReceptora': codigoPlantaReceptora
        }
        URL = os.getenv("PLANTA_PUT_ENXERTIA")
        Editar_Informacao(params, URL) 

async def Adicionar_planta():
    URL = os.getenv("AUTOMACAO_PLANTA_POST")
    tabela = openpyxl.load_workbook("Planta.xlsx")
    tabela_dados = tabela['Plan1']
    for linha in tabela_dados.iter_rows(min_row=2):
        nome_cientifico = linha[0].value if linha[0].value else ""
        nome_popular = linha[1].value if linha[1].value else ""
        instrucoes = linha[2].value if linha[2].value else ""
        cavalo = linha[3].value
        quantidade = int(linha[4].value) if linha[4].value else 0
        if cavalo == "False":
            cavalo = False
        else:
            cavalo = True
        params = {
            'nomeCientifico': nome_cientifico,
            'nomePopular': nome_popular,
            'instrucoes': instrucoes,
            'cavalo': cavalo,
            'quantidade': quantidade
        }
        Enviar_Informacao(params=params,url= URL)
    #tabela .save("Planta.xlsx")

#métodos de menus de navegação  

def menu():
    print("1 - Área de Plantio")
    print("2 - Subárea de Plantio")
    print("3 - Planta")
    print("4 - Alterar Ciclo")
    print("5 - Sair")
    opcao = int(input("Escolha uma opção: "))

    if opcao == 1:
        submenu_Area_Plantio()
    elif opcao == 2:
        submenu_Subarea_Plantio()
    elif opcao == 3:
        submenu_Planta()
    elif opcao == 4:
        asyncio.run(Alterar_Ciclo())
    elif opcao == 5:
        print("Saindo...")
        exit()
    else:
        print("Opção inválida")
        menu()  

def submenu_Planta():
    print("1 - Adicionar")
    print("2 - Enxertia")
    print("3 - Buscar Por Código")
    print("4 - Listar Todas")
    print("5 - Listar Germinações")
    print("6 - Listar Mudas")
    print("7 - Listar Crescimento")
    print("8 - Listar Floração")
    print("9 - Listar Frutificação")
    print("10 - Listar Maturaçao")
    print("11 - Listar Fim Ciclo")
    print("12 - Retornar ao Menu anterior")
    opcao = int(input("Escolha uma opção: "))
    if opcao == 1:
        asyncio.run(Adicionar_planta())
    elif opcao == 2:
        asyncio.run(RealizarEnxertia()) 
    elif opcao == 3:
        url = os.getenv("PLANTA_GET_CODIGO")
        codigo = input("Código a ser localizado: ")
        asyncio.run(Buscar_Informação(url=url,param=codigo))
    elif opcao == 4:
        url = os.getenv("PLANTA_GET_LIST")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 5:
        url = os.getenv("PLANTA_GET_LIST_GERMINACAO")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 6:
        url = os.getenv("PLANTA_GET_LIST_MUDA")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 7:
        url = os.getenv("PLANTA_GET_LIST_CRESCIMENTO")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 8:
        url = os.getenv("PLANTA_GET_LIST_FLORACAO")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 9:
        url = os.getenv("PLANTA_GET_LIST_FRUTIFICACAO")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 10:
        url = os.getenv("PLANTA_GET_LIST_MATURACAO")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 11:
        url = os.getenv("PLANTA_GET_LIST_FIM")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 12:
        menu()
    else:
        print("Opção Inválida")
        menu()


def submenu_Area_Plantio():
    print("1 - Adicionar")
    print("2 - Buscar por nome")
    print("3 - Listar")
    print("4 - Adubação Geral")
    print("5 - Retornar ao menu anterior")
    opcao = int(input("Escolha uma opção: "))
    if opcao == 1:
        asyncio.run(Adicionar_Area_Palntio())
    elif opcao == 2:
        url = os.getenv("AREA_PLANTIO_GET")
        NomeBusca = input("Nome Identificador: ")
        asyncio.run(Buscar_Informação(url=url, param=NomeBusca))
    elif opcao == 3:
        url = os.getenv("AREA_PLANTIO_GET_LIST")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 4:
        asyncio.run(Adicionar_Adubacao_Geral())
    elif opcao == 5:
        menu()
    else:
        print("Opção Inválida")
        menu()

def submenu_Subarea_Plantio():
    print("1 - Adicionar")
    print("2 - Buscar por nome")
    print("3 - Listar")
    print("4 - Adubação Individual")
    print("5 - Retornar ao menu anterior")
    opcao = int(input("Escolha uma opção: "))
    if opcao == 1:
        asyncio.run(Adicionar_Subarea_Palntio())
    elif opcao == 2:
        url = os.getenv("SUBAREA_GET")
        numeroBusca = input("Número Identificador: ")
        asyncio.run(Buscar_Informação(url=url,param=numeroBusca))
    elif opcao == 3:
        url = os.getenv("SUBAREA_GET_LIST")
        asyncio.run(Buscar_Informação_Lista(url=url))
    elif opcao == 4:
        asyncio.run(Adicionar_Adubacao_Individual())
    elif opcao == 5:
        menu()
    else:
        print("Opção Inválida")
        menu()
#métodos de CRUD

def Editar_Informacao(params, url):
    headers = {
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}
    params = params
    try:
        resposta = requests.put(
        url=url,
        json=params,
        headers=headers,
        params=params,
        timeout=30  # Timeout em segundos
        )
        
        if resposta.status_code == 200:
            print(f"Dados enviados com sucesso!")
        else:
            print(f"Erro ao enviar dados Status code: {resposta.status_code}")
    except Exception as e:
        print(f"Erro ao fazer a requisição.")

def Enviar_Informacao(params, url):
    headers = {
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}
    params = params
    try:
        resposta = requests.post(
        url=url,
        json=params,
        headers=headers,
        params=params,
        timeout=30  # Timeout em segundos
        )
        
        if resposta.status_code == 201:
            print(f"Dados enviados com sucesso!")
        else:
            print(f"Erro ao enviar dados Status code: {resposta.status_code}")
    except Exception as e:
        print(f"Erro ao fazer a requisição.")

def Buscar_Informação_Lista(url):
    try:
        response = requests.get(url=url)
        response.raise_for_status()
        data = response.json()
        
        if isinstance(data, list):
            for item in data:
                print("\n--- Item ---")
                for key, value in item.items():
                    print(f"{key}: {value}")
        else:
            print("\n--- Response ---")
            for key, value in data.items():
                print(f"{key}: {value}")
                
    except requests.exceptions.RequestException as e:
        return ""
    
def Buscar_Informação(url, param):
    if(url and param):
        try:
            params = {'codigo': param}
            response = requests.get(url=url, params=params)
            response.raise_for_status()
            data = response.json()
            
            print("\n--- Response ---")
            if isinstance(data, list):
                for item in data:
                    print("\n--- Item ---")
                    for key, value in item.items():
                        print(f"{key}: {value}")
            else:
                for key, value in data.items():
                    print(f"{key}: {value}")
                    
            return data
        except requests.exceptions.RequestException as e:
            return ""


if __name__ == '__main__':
    while True:
        try:
            menu()
        except Exception as e:
            print(f"Ocorreu um erro: {e}") 

 